import os
import mysql.connector as con
from openpyxl import load_workbook

def exToTable(filename):

    wb = load_workbook(filename = '.\\excel\\' + filename + '.xlsx')
    sheet = wb['Sheet1']

    connection = con.connect(
        host='localhost',
        user='george',
        password='root',
        database='sp'
    )

    cur = connection.cursor()

    titles = []

    for column in sheet.iter_cols():
        titles.append(column[0].value.replace(" ", "_"))

    query = f'create table if not exists {filename.replace("-","_")} ({titles[0]} DATE NOT NULL, {titles[1]} VARCHAR(30), {titles[2]} VARCHAR(30), {titles[3]} VARCHAR(30), {titles[4]} VARCHAR(30), {titles[5]} VARCHAR(30), {titles[6]} INT, {titles[7]} VARCHAR(50))'

    cur.execute(query)
